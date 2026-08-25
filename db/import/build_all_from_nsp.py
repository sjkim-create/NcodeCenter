# 2_New_NSP_Ncode_List.xlsx → 코드 프로젝트 seed + 편집 상세(책 단위, 심볼 모드별) 데이터
import sys, os, re, json
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.join(HERE, "..", "..", "web", "data")
# 원본 엑셀: db/source 우선, 없으면 Downloads
# 편집 현황(= 편집 관리 원장). 소리펜/필기펜 목록은 코드 할당 자료라 여기 실린 교재는 모두 "편집"으로 본다.
NAME = "2_New_NSP_Ncode_List.xlsx"
EDIT_NAME = "[데이터확인용]NcodeCenter_편집현황_실사용아님.xlsx"
_cands = [os.path.join(HERE, "..", "source", EDIT_NAME), os.path.join(HERE, "..", "source", NAME),
          os.path.join(r"C:\Users\NeoLab\Downloads", NAME)]
SRC = next((p for p in _cands if os.path.exists(p)), _cands[0])
print("원본:", os.path.abspath(SRC))

SOUND = [("기본",13),("멀티터치·번역",14),("슬롯전환",15),("전체듣기",16),("게임",17),("프롬프트",18),("발음평가",19)]
PEN = [("필기펜 기본",20),("캘린더",21),("링크",22),("교원구몬/KEP",23)]
TOT_I, METHOD_I, TYPE_I = 24, 25, 33

def num(v):
    try:
        if v in (None,"","-"): return 0.0
        return float(v)
    except: return 0.0
def dstr(v):
    try: return v.date().isoformat()
    except: return ""

wb = load_workbook(SRC, data_only=True)
companies=[]; projects=[]; logs=[]; edit_customers=[]
G_sound={k:0 for k,_ in SOUND}; G_pen={k:0 for k,_ in PEN}
T_books=T_pages=T_sym=T_size=0
cid=pid=lid=0
comp_by_name={}   # 같은 고객사명은 1개 회사로 통합 (owner별로는 프로젝트로 분리)

for ws in wb.worksheets:
    m=re.search(r"\((\d+)\)", ws.title)
    owner=int(m.group(1)) if m else 0
    name=re.sub(r"\(\d+\)","",ws.title).strip()
    books=[]; pages=size=sym=withsym=0
    smode={k:0 for k,_ in SOUND}; pmode={k:0 for k,_ in PEN}
    kinds=set(); methods={}; groups={}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or r[0] in (None,""): continue
        sec=int(num(r[2])); own=int(num(r[3])); bk=int(num(r[4]))
        pg=int(num(r[6])); sz=num(r[12])
        kind="N" if "PDS3" in str(r[1] or "") else "G"
        ty=str(r[TYPE_I] or "").strip() if len(r)>TYPE_I else ""
        method=str(r[METHOD_I]).strip() if len(r)>METHOD_I and r[METHOD_I] else ""
        sm=[int(num(r[i])) if len(r)>i else 0 for _,i in SOUND]
        pm=[int(num(r[i])) if len(r)>i else 0 for _,i in PEN]
        total_xlsx=int(num(r[TOT_I])) if len(r)>TOT_I else 0
        # 세부 모드가 비었지만 총합이 있으면 펜 타입 기준 기본에 배치
        if sum(sm)+sum(pm)==0 and total_xlsx>0:
            if "필기펜" in ty: pm[0]=total_xlsx
            else: sm[0]=total_xlsx
        bsym=sum(sm)+sum(pm)
        pages+=pg; size+=sz; sym+=bsym
        if bsym>0: withsym+=1
        kinds.add(kind)
        if method:
            for one in [x.strip() for x in method.split(",") if x.strip()]:
                methods[one]=methods.get(one,0)+1
        for (k,_),v in zip(SOUND,sm): smode[k]+=v; G_sound[k]+=v
        for (k,_),v in zip(PEN,pm): pmode[k]+=v; G_pen[k]+=v
        g=groups.setdefault((sec,own), {"bmin":bk,"bmax":bk,"pages":0,"date":dstr(r[7])})
        g["bmin"]=min(g["bmin"],bk); g["bmax"]=max(g["bmax"],bk); g["pages"]+=pg
        if dstr(r[7]) and (not g["date"] or dstr(r[7])<g["date"]): g["date"]=dstr(r[7])
        book={"b":bk,"s":sec,"o":own,"k":kind,"pg":pg,"t":str(r[9] or ""),
              "f":str(r[10] or ""),"bytes":int(sz),"ty":ty or "소리펜",
              "sm":sm,"pm":pm,"m":method,"d":dstr(r[7])}
        def sval(idx):
            v=r[idx] if len(r)>idx else None
            return "" if v in (None,"","-") else str(v).strip()
        sp=int(num(r[5])) if len(r)>5 and num(r[5])>0 else 0
        if sp: book["sp"]=sp
        setc=int(num(r[29])) if len(r)>29 and num(r[29])>0 else 0
        if setc: book["set"]=setc
        for k2,idx,dt in [("del",8,1),("nmod",11,1),("iss",26,0),("use",28,0),
                          ("pmdl",34,0)]:
            val=(dstr(r[idx]) if len(r)>idx else "") if dt else sval(idx)
            if val: book[k2]=val
        # 세부내역·출력용파일·APP데이터 → 첨부(링크 기본 + 설명)
        for k2,idx in [("det",30),("out",31),("app",32)]:
            v=sval(idx)
            if v: book[k2]={"mode":"link","value":v,"note":""}
        # 메모1~3 → 업무요청 메모(logs) 로 이전 (kind=메모)
        blogs=[]
        for idx in (35,36,37):
            v=sval(idx)
            if v: blogs.append({"id":len(blogs)+1,"no":len(blogs)+1,"kind":"메모","content":v,"date":dstr(r[7]),"author":""})
        if blogs: book["logs"]=blogs
        books.append(book)
    if name in comp_by_name:
        cur_cid = comp_by_name[name]
    else:
        cid += 1; cur_cid = cid; comp_by_name[name] = cur_cid
        companies.append({"id":cur_cid,"name":name,"manager":"","contact":"","address":"",
                          "bizNo":"","bankName":"","accountNo":"","docs":[]})
    pid+=1
    issued=[]
    for bi,((sec,own),g) in enumerate(sorted(groups.items())[:10]):
        issued.append({"id":bi+1,"date":g["date"] or "2020-01-01","codes":int(g["pages"]),
                       "section":sec,"owner":own,"bookStart":g["bmin"],"bookEnd":g["bmax"],
                       "pageStart":0,"pageEnd":0})
    projects.append({"id":pid,"name":f"{name} 코드발급 · O{owner}","companyId":cur_cid,
                     "service":"NEONOTE","grade":"","issued":issued,
                     "editingOwner":owner,"symbols":int(sym)})
    lid+=1
    log_no = sum(1 for l in logs if l["companyId"] == cur_cid) + 1   # 고객사 내 안정 번호
    logs.append({"id":1000+lid,"no":log_no,"companyId":cur_cid,"projectId":pid,"date":"2026-03-02",
                 "kind":"메모","content":f"{name} · O{owner} · 발급 {int(pages):,}p · 편집 심볼 {int(sym):,}","author":"이영업"})
    real_owners=sorted({b["o"] for b in books})
    edit_customers.append({"customer":name,"owner":str(owner),"owners":real_owners,"codeKinds":sorted(kinds),
        "books":len(books),"pages":int(pages),"symbols":int(sym),
        "soundSymbols":int(sum(smode.values())),"penSymbols":int(sum(pmode.values())),
        "withSymbolBooks":withsym,"sizeMB":round(size/1e6),
        "soundBreakdown":smode,"penBreakdown":pmode,
        "topMethods":sorted(methods.items(),key=lambda x:-x[1])[:4],
        "bookRows":books})
    T_books+=len(books); T_pages+=pages; T_sym+=sym; T_size+=size

edit_customers.sort(key=lambda c:-c["symbols"])
json.dump({"companies":companies,"projects":projects,"logs":logs},
          open(os.path.join(WEB,"seed-customers.json"),"w",encoding="utf-8"),ensure_ascii=False,indent=1)
json.dump({"summary":{"customers":len(edit_customers),"books":T_books,"pages":int(T_pages),
                      "symbols":int(T_sym),"sizeGB":round(T_size/1e9,1)},
           "grandSound":G_sound,"grandPen":G_pen,"customers":edit_customers},
          open(os.path.join(WEB,"editing-data.json"),"w",encoding="utf-8"),ensure_ascii=False)
print(f"companies={len(companies)} editBooks={T_books} symbols={int(T_sym):,}")
print("소리펜 모드:", G_sound); print("필기펜 모드:", G_pen)
print("editing-data.json", round(os.path.getsize(os.path.join(WEB,"editing-data.json"))/1024),"KB")
