# nkey(HLP).xlsx = N Key 티켓 발급 관리 대장 → web/data/ticket-ledger.json
#   · 'Ncode 티켓 '  : Section 3 (PDS3)  — owner/book범위/프로젝트/구분/발급자/페이지/발급일/기간
#   · 'Scode 티켓'   : Section 44 (Scode) — 동일 레이아웃(오너부터 시작)
#   · '2018 이전 티켓 발급' : 소량, 다른 레이아웃
# 구분(판매용/무상/미사용) → 정산 유형(유료/무료/미정) 기본값으로 매핑. 금액은 미등록(0).
import sys, os, re, json
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "..", "source", "nkey(HLP).xlsx")
WEB = os.path.join(HERE, "..", "..", "web", "data")

def txt(v):
    return "" if v in (None, "", "-") else str(v).strip()
def inum(v):
    m = re.search(r"-?\d+", str(v or "").replace(",", ""))
    return int(m.group()) if m else None
def dstr(v):
    try: return v.date().isoformat()
    except Exception:
        s = txt(v)
        m = re.search(r"(20\d{2})[.\-/\s]+(\d{1,2})[.\-/\s]+(\d{1,2})", s)
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else ""
def period(v):
    s = txt(v)
    if not s or "무제한" in s: return "무제한"
    m = re.search(r"(20\d{2})[.\-/\s]+(\d{1,2})[.\-/\s]+(\d{1,2})", s)
    return f"~{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else s

# 구분 → (정산유형, 비고)
def billing_of(gubun):
    g = gubun.replace(" ", "")
    if "무상" in g or "무료" in g: return "무료", gubun
    if "미사용" in g: return "미정", gubun or "미사용"
    if "판매" in g or "유료" in g: return "유료", gubun          # 금액은 추후 등록
    return "미정", gubun

wb = load_workbook(SRC, data_only=True)
tickets = []
no = 0

def add(sec, pt, owner, bs, be, project, gubun, issuer, pages, date, per):
    global no
    if not project: return
    no += 1
    bill, note = billing_of(gubun)
    be2 = be if be is not None else bs
    # BookEnd 는 저장하지 않고 Book 볼륨(권수)으로 환산해 기록 (현재 발급 화면과 동일)
    bvol = (be2 - bs + 1) if (bs is not None and be2 is not None) else (1 if bs is not None else None)
    valid = "99999999 (무제한)" if per == "무제한" else per.lstrip("~")
    summary = f"{pt} S{sec}/O{owner}" + (f"/B{bs}" if bs is not None else "") + \
              (f" · Book {bvol}권" if bvol else "") + \
              (f" · P0~{pages-1}" if pages else "") + f" · 기간 {per}"
    # 현재 N Key 발급이 기록하는 항목 순서와 맞춘다
    params = {"CompanyName": project, "IssuedTime": (date or "").replace("-", "") or "-",
              "ValidUntilTime": valid, "Section": sec, "Owner": owner, "TicketVersion": 1}
    if bs is not None: params["BookStart"] = bs
    if bvol: params["BookVolume"] = bvol
    params["PageStart"] = 0
    if pages: params["PageVolume"] = pages
    params["PatternType"] = pt
    params["TicketType"] = "Unlimited"
    params["SeparateEachBook"] = "N (1개 티켓 병합)"
    if gubun: params["구분(대장)"] = gubun
    if issuer: params["담당(대장)"] = issuer
    tickets.append({
        "src": "ledger", "no": no, "kind": "N", "company": project,
        "at": (date or "2020-01-01") + "T00:00:00", "by": issuer, "summary": summary,
        "params": params, "billing": bill, "amount": 0,
        "billNote": note, "period": per,
    })

# ── Ncode 티켓 (Section 3) ──  col1=owner 2=bs 3=be 4=project 5=구분 6=발급자 7=page 8=date 9=기간
ws = wb["Ncode 티켓 "]
for row in ws.iter_rows(min_row=2, values_only=True):
    g = lambda i: row[i] if len(row) > i else None
    project = txt(g(4))
    if not project: continue
    add(3, "Ncode_PDS3", inum(g(1)), inum(g(2)), inum(g(3)), project,
        txt(g(5)), txt(g(6)), inum(g(7)), dstr(g(8)), period(g(9)))

# ── Scode 티켓 (Section 44) ── col0=owner 1=bs 2=be 3=project 4=구분 5=발급자 6=page 7=date 8=기간
ws = wb["Scode 티켓"]
for row in ws.iter_rows(min_row=3, values_only=True):
    g = lambda i: row[i] if len(row) > i else None
    project = txt(g(3))
    if not project or project.lower() in ("프로젝트명",): continue
    add(44, "Scode", inum(g(0)), inum(g(1)), inum(g(2)), project,
        txt(g(4)), txt(g(5)), inum(g(6)), dstr(g(7)), period(g(8)))

# ── 2018 이전 (소량) ── 섹션/오너/북코드/페이지/회사명/메모/기타/티켓기한/발급일자/발급자
ws = wb["2018 이전 티켓 발급"]
for row in ws.iter_rows(min_row=2, values_only=True):
    g = lambda i: row[i] if len(row) > i else None
    company = txt(g(4))
    sec = inum(g(0))
    if not company or sec is None: continue
    bs, be = None, None
    bm = re.findall(r"\d+", str(g(2) or ""))
    if bm: bs = int(bm[0]); be = int(bm[1]) if len(bm) > 1 else int(bm[0])
    add(sec, "Ncode_PDS3", inum(g(1)), bs, be, company,
        "", txt(g(9)), inum(g(3)), dstr(g(8)), period(g(7)))

json.dump(tickets, open(os.path.join(WEB, "ticket-ledger.json"), "w", encoding="utf-8"),
          ensure_ascii=False, separators=(",", ":"))

from collections import Counter
bc = Counter(t["billing"] for t in tickets)
print(f"ticket-ledger.json · {len(tickets)}건 · 정산기본값 {dict(bc)}")
print("고객사:", ", ".join(sorted({t['company'] for t in tickets})))
