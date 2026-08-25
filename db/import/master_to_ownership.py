#!/usr/bin/env python3
# NcodeCenter — 마스터 시트("오너코드_발급리스트") → 소유권 레코드 파싱/검증
# 사용: python db/import/master_to_ownership.py "<xlsx>" [--json]
# 마스터 = ACCOUNT × Section × Owner × 북코드범위(CONTENTS) × 페이지범위(PAGE).
#   헤더 r2: No, ACCOUNT, Section, Owner, CONTENTS, PAGE, X, Y, REG DATE  (데이터 r3+)
import sys, re, json
from collections import Counter
from openpyxl import load_workbook
try: sys.stdout.reconfigure(encoding="utf-8")
except Exception: pass

# N코드 정보표 (정합 확인용): 정보표에 있는 Section만
DIM = {"PDS3":{0,3,5,10,11,14,15}, "PDS2":{0,3,14}}
KNOWN_SECTIONS = DIM["PDS3"] | DIM["PDS2"]

def rng(s):
    if s is None: return None
    m = re.search(r"(\d[\d,]*)\s*~\s*(\d[\d,]*)", str(s))
    if m: return (int(m.group(1).replace(",","")), int(m.group(2).replace(",","")))
    m2 = re.fullmatch(r"\s*(\d[\d,]*)\s*", str(s))
    if m2: v=int(m2.group(1).replace(",","")); return (v,v)
    return None

def num(v):
    try: return int(float(v))
    except: return None

def main():
    if len(sys.argv) < 2:
        print('사용: python master_to_ownership.py "<xlsx>" [--json]'); sys.exit(2)
    wb = load_workbook(sys.argv[1], data_only=True, read_only=True)
    if "오너코드_발급리스트" not in wb.sheetnames:
        print("마스터 시트('오너코드_발급리스트') 없음"); sys.exit(1)
    ws = wb["오너코드_발급리스트"]
    rows = list(ws.iter_rows(values_only=True))

    recs=[]; bad=0
    for r in rows[3:]:
        if not r or r[1] in (None,""): continue
        sec=num(r[2]); own=num(r[3])
        if sec is None or own is None: bad+=1; continue
        recs.append(dict(account=str(r[1]).strip(), section=sec, owner=own,
                         books=rng(r[4]), pages=rng(r[5]),
                         book_raw=str(r[4]) if r[4] is not None else "",
                         page_raw=str(r[5]) if r[5] is not None else "",
                         reg=str(r[8]) if len(r)>8 and r[8] is not None else ""))
    accounts=sorted({x["account"] for x in recs})
    secs=sorted({x["section"] for x in recs})
    unknown=[s for s in secs if s not in KNOWN_SECTIONS]

    print(f"소유권 레코드: {len(recs)} (파싱불가 {bad})")
    print(f"ACCOUNT 수: {len(accounts)} · Section 종류: {secs}")
    if unknown: print(f"⚠ 정보표에 없는 Section: {unknown} (레거시/확인 필요)")
    print(f"⚠ 마스터에 product(PDS2/3) 컬럼 없음 → 제품 구분 규칙 필요")
    print(f"\n[Section별 소유 레코드]")
    bs=Counter(x["section"] for x in recs)
    for s in secs: print(f"  S{s}: {bs[s]}" + ("  ← 정보표 없음" if s in unknown else ""))
    print(f"\n[ACCOUNT 상위 12]")
    for a,c in Counter(x["account"] for x in recs).most_common(12): print(f"  {a:24} {c}")

    if "--json" in sys.argv:
        with open("master_ownership.json","w",encoding="utf-8") as f:
            json.dump(recs, f, ensure_ascii=False, indent=1)
        print("\n→ master_ownership.json 저장")

if __name__ == "__main__":
    main()
