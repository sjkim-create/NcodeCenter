# 소리펜(NSP) + 필기펜(NWP) + 편집현황 3개 원본을 모두 읽어
#   web/data/seed-customers.json (코드 프로젝트) + editing-data.json (편집 프로젝트) 생성
#
# 규칙(사용자 확인):
#  - 각 시트 = 고객사 1건. 단 아래 시트는 제외
#      NSP: Ncode_info / NWP: 오너코드_발급리스트, 티켓발급리스트, P2UI Param
#  - 코드 할당 원본 = 소리펜/필기펜 목록,  편집 여부 원장 = 편집현황 파일
#    → 편집현황에 실린 교재만 "편집"으로 표시(ed=True)
#  - 시트마다 레이아웃이 달라 "헤더명 기반"으로 컬럼을 찾는다.
import sys, os, re, json
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "..", "source")
WEB = os.path.join(HERE, "..", "..", "web", "data")

NWP_FILE = "(필기펜)NWP_Ncode_List.xlsx"
# 고객사(원장) 구성은 소리펜(NSP)·필기펜(NWP) 두 파일의 시트명이 정본이다.
FILES = [
    ("sound", "(소리펜)NSP_Ncode_List.xlsx", {"ncode_info", "게임펌웨어엔진"}),
    ("pen", NWP_FILE, {"오너코드_발급리스트", "티켓발급리스트", "p2ui param"}),
]
# 편집현황 파일은 시트명(고객사)이 NSP/NWP 원장과 어긋나 있어 고객사 식별에 쓰지 않는다.
# 대신 코드(k/s/o/b)로 매칭해서 '편집 여부(ed)'와 '심볼수(sm/pm)'만 원장 책에 이식한다.
EDIT_FILE = "[데이터확인용]NcodeCenter_편집현황_실사용아님.xlsx"

# ── 커먼 코드(공유 코드) 시트 ──────────────────────────
# 여러 고객사가 함께 쓰는 코드를 한 시트에서 관리한다. 시트의 "고객사" 열이 실제 사용 업체이므로
#   · 시트 자체는 코드 보유자(holder) 한 곳으로 묶고
#   · 행의 고객사는 cu(사용 고객사)로 남기며, 고객사 관리에 NSP/NWP 커먼 코드 체크로 등록한다.
COMMON_SHEETS = {
    "Common (21)":                     ("NSP", "Common-21"),      # 옛 레퍼런스(21)
    "Common 추가 언어 슬롯 (964~983)": ("NSP", "Common-964"),     # 옛 레퍼런스(964~983)
    "네오노트 (27)":                   ("NWP", "네오노트-27"),
    "네오노트 (1012)":                 ("NWP", "네오노트-1012"),
    "스마트클래스키트 (1013)":         ("NWP", "스마트클래스키트-1013"),  # 옛 S3_O1013_PUI
}
# ── 고객사명 표기 교정 ─────────────────────────────────────
# 소스 시트명이 바뀌었으나 파일(이미지·도형 포함)에 아직 반영되지 않은 건 여기서 교정한다.
# 키는 nz(정규화: 소문자·공백제거)된 원본 시트명, 값은 표시할 고객사명.
CUST_RENAME = {"구몬c": "구몬S"}   # 구몬C(1006) → 구몬S(1006)

# 프로젝트 종료(사업 정리) 고객사 — 코드 발급 이력만 유지, 화면에서 비활성 표시. {nz(고객사명): 종료 메모}
CLOSED_ACCOUNTS = {"humedical-28": "사업 정리로 프로젝트 종료 · 동일 코드(S3/O28)를 엠베스트-28 가 이어서 사용"}

# ── 언어 슬롯 매핑 (PDS2 · Section 3 공유코드) ────────────────────────
# COMMON-21(기본 언어) + 확장 언어 슬롯(964~983)을 언어에 매핑한다. 등록 UI 없이 여기서 직접 관리.
# 언어 확장 시 이 표에 (owner, slot, 언어) 한 줄만 추가하면 된다. (owner=공유코드 owner, slot=전역 언어 슬롯 번호)
# casterN: 21 기본 언어로 편집하다 언어 확장하면 해당 언어의 확장 슬롯 owner 코드가 함께 사용된다(예: 러시아어→964).
LANG_SLOTS_PDS, LANG_SLOTS_SECTION, LANG_BASE_OWNER = "G", 3, 21   # G=PDS2(소리펜)
LANG_SLOTS = [
    # (owner, slot, 언어) — 기본 언어 (Common-21)
    (21, 1, "한국어"), (21, 2, "영어"), (21, 3, "중국어"), (21, 4, "일본어"),
    # 확장 언어 슬롯 (964~983) — 나머지는 비어 있음, 언제든 확장 가능
    (964, 5, "베트남어"), (964, 6, "러시아어"), (964, 7, "몽골어"), (964, 8, "캄보디아어(크메르어)"),
    (965, 9, "스리랑카어"), (965, 10, "필리핀어(따갈로그어)"),
]

# ── (미리보기) 오너별 고객사 분리 ─────────────────────────────
# "1 고객사 = 1 OWNER" 모델 시험. 여기 든 고객사(nz 기준)는 보유 오너마다 "{이름}-{오너}"로 쪼갠다.
# 대장 재정리 후 전면 적용 예정 — 지금은 구몬만 시험.
SPLIT_BY_OWNER = {"구몬"}

# 대장에서 ACCOUNT 병합셀 '연속행'이라도 오너마다 개별 고객사 "{base}-{owner}"로 쪼갤 계정 (nz 기준).
# 예: Cake 984~1003 은 병합셀 연속행이지만 실제로는 오너별 개별 발급이다 → Cake-984 … Cake-1003.
SPLIT_EACH_OWNER = {"cake"}

# 편집 프로젝트 좌측 목록에서 이 대장 고객사(공통코드 홀더)를 코드(타입·섹션)별로 쪼갠다.
#   companies·ownership 은 그대로(네오노트-27) 두고, edit_customers(편집 프로젝트)만 분리한다.
#   예: 네오노트-27 → 네오노트-0-27(N/S0/O27) · 네오노트-3-27(N/S3/O27) · 네오노트-IDS-27(A/S4/O27) …
COMMON_SPLIT = {"네오노트-27"}
def split_customer_name(name, k, s):
    if name not in COMMON_SPLIT: return name
    base, _, own = name.rpartition("-")            # "네오노트-27" → ("네오노트","-","27")
    return f"{base}-IDS-{own}" if k == "A" else f"{base}-{s}-{own}"

SOUND_N, PEN_N = 7, 4
SOUND_I = [13, 14, 15, 16, 17, 18, 19]   # 심볼 블록(편집현황 표준 레이아웃)
PEN_I = [20, 21, 22, 23]
TOT_I = 24

def num(v):
    try:
        if v in (None, "", "-"): return 0.0
        return float(v)
    except Exception: return 0.0
def dstr(v):
    try: return v.date().isoformat()
    except Exception: return ""
def txt(v):
    return "" if v in (None, "", "-") else str(v).strip()
def first_int(v, default=0):
    m = re.search(r"\d+", str(v or ""))
    return int(m.group()) if m else default
def nz(x): return re.sub(r"\s+", "", str(x or "")).lower()

LABELS = {
    "d": ["북코드발급일자", "발급일자"],
    "del": ["북코드삭제일자", "삭제일자"],
    "nmod": ["ncp2최종수정날짜", "수정한날짜", "수정날짜"],
    "bytes": ["ncp2파일크기", "파일크기"],
    "sym": ["매핑개수", "소리펜심볼"],
    "pg": ["pdf페이지수", "페이지수", "totalpage"],
    "sp": ["startpage", "시작페이지"],
    "use": ["편집완료여부", "사용여부"],
    "book": ["북코드", "book"],
    "sec": ["섹션코드", "section"],
    "own": ["오너코드", "오너", "owner"],
    "title": ["프로젝트명", "교재명", "컨텐츠명", "콘텐츠명", "교재"],
    "file": ["ncp2파일명", "ncp파일명", "nproj파일명", "zip파일명", "파일명"],
    "m": ["편집방식"],
    "memo": ["기타/메모", "기타", "메모"],
    "cu": ["고객사명", "고객사"],           # 커먼 코드 시트의 실제 사용 업체
    "edok": ["편집"],                       # 레퍼런스 시트 '편집' O/X (실제 편집 여부)
    "usefor": ["데모용/전시용/공식제품용"],  # 용도
    "pmdl": ["모델구분"],
    "kind": ["코드구분", "pds/ids", "구분"],
}
# 긴 라벨 우선 매칭 (예: '북코드발급일자' 가 '북코드' 보다 먼저)
FLAT = sorted(((n, f) for f, ns in LABELS.items() for n in ns), key=lambda x: -len(x[0]))

def top_block_so(ws, scan=6):
    """상단 블록의 Section / Owner 값 (예: r2 'Section | 3.0')"""
    sec = own = None
    for row in ws.iter_rows(min_row=1, max_row=scan, values_only=True):
        if not row: continue
        for i, c in enumerate(row):
            lab = nz(c)
            if lab in ("section", "섹션") and sec is None:
                for v in row[i + 1:]:
                    if txt(v): sec = first_int(v); break
            if lab in ("owner", "오너") and own is None:
                for v in row[i + 1:]:
                    if txt(v): own = first_int(v); break
    return sec, own

def find_header(ws, scan=14):
    """헤더 행과 컬럼 매핑 탐색 → (row_idx, colmap, has_symbol_block)"""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True), 1):
        if not row: continue
        cells = [nz(c) for c in row]
        cmap = {}
        for ci, cell in enumerate(cells):
            if not cell: continue
            for name, field in FLAT:                 # 긴 라벨부터
                if field in cmap: continue
                if cell == name or cell.startswith(name):
                    cmap[field] = ci; break
        if "book" in cmap and ("title" in cmap or "file" in cmap) and len(cmap) >= 3:
            return i, cmap, any(c == "기본" for c in cells)
    return None, None, False

# 커먼 코드 시트의 "고객사" 칸은 손입력이라 상태 메모가 섞여 있다 → 업체명이 아닌 값 걸러내기
CU_NOISE = {"-", "미정", "미상", "테스트", "test", "상용", "시범설치용", "산학연구데모", "b2b&b2c",
            "casestudy", "데모", "demo", "샘플", "sample", "공용", "기타", "없음", "n/a", "na"}
def clean_cu(v):
    s = re.sub(r"\s+", " ", str(v or "").replace("\n", " ")).strip(" ·,/")
    if not s or s in ("-",): return ""
    if nz(s) in CU_NOISE: return ""
    if re.fullmatch(r"[\d.,~\-]+", s): return ""       # 숫자/범위만 적힌 칸
    if len(s) < 2 or len(s) > 40: return ""
    return s

def book_range(v):
    """'3 ~ 22' / '0~19' / '5' → (시작, 끝)"""
    m = re.findall(r"\d+", str(v or "").replace(",", ""))
    if not m: return (0, 0)
    a = int(m[0])
    return (a, int(m[1]) if len(m) > 1 else a)

def parse_sheet(ws, default_ty, s_hint=None, o_hint=None, common=False):
    hrow, cm, has_sym = find_header(ws)
    if not cm: return []
    tb_s, tb_o = top_block_so(ws)                    # 상단 블록의 Section/Owner
    books = []
    last_s = tb_s if tb_s is not None else s_hint
    last_o = tb_o if tb_o is not None else o_hint
    last_k = None
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        if not row: continue
        g = lambda f: (row[cm[f]] if (f in cm and len(row) > cm[f]) else None)

        kind_raw = str(g("kind") or "").upper()
        if "PDS" in kind_raw or "IDS" in kind_raw or "OID" in kind_raw:
            # 코드 종류(좌표 속성): PDS2→G(Gcode) · PDS3→N(Ncode) · IDS→A(초창기, 이력전용)
            #   OID→O — 인덱스 전용(외부 코드 판독용). 같은 S/O 를 다른 종류와 공유하고 B/P 로 구분한다.
            #   ※ Section 44 = PDS4(S-code) 는 좌표(섹션)로 판별하므로 여기서 태깅하지 않는다.
            k = ("O" if "OID" in kind_raw else
                 "G" if "PDS2" in kind_raw else
                 "A" if "IDS" in kind_raw else "N")
            last_k = k
        else:
            if not (txt(g("title")) or txt(g("file"))): continue
            k = last_k or "N"

        sec_cell, own_cell = g("sec"), g("own")
        sec = int(num(sec_cell)) if txt(sec_cell) != "" else (last_s if last_s is not None else 0)
        own = int(num(own_cell)) if txt(own_cell) != "" else (last_o if last_o is not None else 0)
        last_s, last_o = sec, own

        bs, be = book_range(g("book"))
        has_book = bool(re.search(r"\d", str(g("book") or "")))   # 북코드 셀에 실제 숫자가 있나
        b = {"b": bs, "s": sec, "o": own, "k": k,
             "pg": int(num(g("pg"))), "t": txt(g("title")), "f": txt(g("file")),
             "bytes": int(num(g("bytes"))), "ty": default_ty,
             "sm": [0] * SOUND_N, "pm": [0] * PEN_N,
             "m": txt(g("m")), "d": dstr(g("d"))}
        if not has_book: b["nb"] = 1                    # 북코드 없는 행(중복/변형) → 실제 book 아님
        if be > bs: b["b2"] = be                       # '3 ~ 22' 처럼 범위로 발급된 행
        if common:                                     # 커먼 코드 시트 → 행마다 실제 사용 고객사
            cu = clean_cu(g("cu"))
            if cu: b["cu"] = cu
            # 편집 여부: 공통코드는 '프로젝트명(교재명)'이 있으면 편집(실제 작업), 없으면 사용가능(코드만 확보).
            #   일부 시트에 '편집' O/X 컬럼이 있으면 그것도 편집으로 인정.
            edok = txt(g("edok")).upper()               # (있을 때만) 편집 O/X
            b["ea"] = 1 if (b["t"] or edok in ("O", "0", "Y", "예", "완료")) else 0
        uf = txt(g("usefor"))
        if uf and uf != "-": b["uf"] = uf
        # 시작 페이지: 셀에 값이 있으면(0 포함) 저장 → 0/1/2 구분 (엠베스트 등)
        if txt(g("sp")) != "": b["sp"] = int(num(g("sp")))
        for f in ("nmod", "del"):
            v = dstr(g(f))
            if v: b[f] = v
        pmdl = txt(g("pmdl"))
        if pmdl: b["pmdl"] = pmdl

        if "sym" in cm:
            b["sm"][0] = int(num(g("sym")))
        elif has_sym:
            sm = [int(num(row[i])) if len(row) > i else 0 for i in SOUND_I]
            pm = [int(num(row[i])) if len(row) > i else 0 for i in PEN_I]
            total = int(num(row[TOT_I])) if len(row) > TOT_I else 0
            if sum(sm) + sum(pm) == 0 and total > 0:
                if "필기펜" in default_ty: pm[0] = total
                else: sm[0] = total
            b["sm"], b["pm"] = sm, pm
            for key, idx in [("iss", 26), ("use", 28), ("pmdl", 34)]:
                v = txt(row[idx]) if len(row) > idx else ""
                if v: b[key] = v
            setc = int(num(row[29])) if len(row) > 29 else 0
            if setc: b["set"] = setc
            for key, idx in [("det", 30), ("out", 31), ("app", 32)]:
                v = txt(row[idx]) if len(row) > idx else ""
                if v: b[key] = {"mode": "link", "value": v, "note": ""}
            ty = txt(row[33]) if len(row) > 33 else ""
            if ty: b["ty"] = ty
            lg = []
            for idx in (35, 36, 37):
                v = txt(row[idx]) if len(row) > idx else ""
                if v: lg.append({"id": len(lg) + 1, "no": len(lg) + 1, "kind": "메모", "content": v, "date": b["d"], "author": ""})
            if lg: b["logs"] = lg

        memo = txt(g("memo"))
        if memo and "logs" not in b:
            b["logs"] = [{"id": 1, "no": 1, "kind": "메모", "content": memo, "date": b["d"], "author": ""}]

        if not (b["t"] or b["f"] or b["pg"] or b["bytes"]
                or (txt(g("book")) and ("PDS" in kind_raw or "IDS" in kind_raw or "OID" in kind_raw))): continue
        books.append(b)
    return books

def sheet_customer(title):
    t = title.strip()
    m = re.match(r"^S([\d,~]+)_O([\d,~]+)_(.+)$", t)     # 필기펜: S3_O54~56_포스트매스
    if m: return m.group(3).strip(), first_int(m.group(1)), first_int(m.group(2))
    m = re.match(r"^(.*?)\(([\d~,]+)\)\s*$", t)           # 소리펜/편집현황: 한솔교육(25)
    if m: return m.group(1).strip(), None, first_int(m.group(2))
    return t, None, None


# ── 오너코드_발급리스트 = 실제 할당 원장 (ACCOUNT/Section/Owner/CONTENTS(book)/PAGE) ──
def parse_alloc_list():
    path = os.path.join(SRC, NWP_FILE)
    if not os.path.exists(path): return []
    wb = load_workbook(path, data_only=True)
    if "오너코드_발급리스트" not in wb.sheetnames: return []
    ws = wb["오너코드_발급리스트"]
    def rng(v):
        m = re.findall(r"[\d,]+", str(v or "").replace(" ", ""))
        if not m: return (None, None)
        a = int(m[0].replace(",", ""))
        b = int(m[1].replace(",", "")) if len(m) > 1 else a
        return (a, b)
    # ACCOUNT(B열) 세로 병합 → 병합 하위 행만 상위 고객사의 '연속(추가) 할당'.
    #   병합 아닌 빈 ACCOUNT 행은 예약/미배정이므로 제외한다. (셀 병합 기준으로만 이어받기)
    from openpyxl.utils import range_boundaries
    acct_top = {}                    # 병합 하위 행 → 상위(값 있는) 행 번호
    for m in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(str(m))
        if c1 == 2 and c2 == 2:
            for rr in range(r1 + 1, r2 + 1): acct_top[rr] = r1
    cv = lambda row, col: ws.cell(row=row, column=col).value
    out = []
    for ridx, r in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not r: continue
        raw = txt(r[1])
        if raw.lower() in ("account",): continue
        if txt(r[3]) == "": continue            # 오너 없는 행 무시
        if raw:                                  # 이름 있는 행 = 새 고객사(할당)
            acct = raw; named = True
            sec = int(num(r[2])) if txt(r[2]) != "" else 0
        elif ridx in acct_top:                   # ACCOUNT 병합 하위 = 위 고객사 연속 할당
            top = acct_top[ridx]
            acct = txt(cv(top, 2)); named = False
            sec = int(num(r[2])) if txt(r[2]) != "" else (int(num(cv(top, 3))) if txt(cv(top, 3)) != "" else 0)
        else:
            continue                             # 병합 아닌 빈 ACCOUNT = 예약/미배정 → 제외
        if not acct: continue
        # 오너 셀에 '~'가 있으면 범위("100~255")로 각 오너를 펼친다(첫 오너만 named). 없으면 단일 오너.
        raw_o = str(r[3])
        if "~" in raw_o:
            onums = re.findall(r"\d+", raw_o.replace(",", ""))
            if not onums: continue
            oa = int(onums[0]); ob = int(onums[1]) if len(onums) > 1 else oa
            if ob < oa: oa, ob = ob, oa
            ob = min(ob, oa + 1023)              # 폭주 방지 상한
        else:
            oa = ob = int(num(r[3]))
        bs, be = rng(r[4]); ps, pe = rng(r[5])
        for i, own in enumerate(range(oa, ob + 1)):
            out.append({"account": acct, "named": (named and i == 0), "section": sec, "owner": own,
                        "book_start": bs, "book_end": be, "page_start": ps, "page_end": pe,
                        "date": dstr(r[8]) if len(r) > 8 else ""})
    return out

ALLOC = parse_alloc_list()

# ── 대장(오너코드_발급리스트) = 고객사 정본 ──────────────────
#   이름 있는 행 = 새 고객사 "{base}-{owner}" (base = 뒤쪽 (숫자) 접미 제거).
#   빈 ACCOUNT 연속행 = 바로 위 고객사의 추가 할당(다른 오너). 예) 교원도요새베트남(4) 아래 O5.
def _base(a): return re.sub(r"\s*\(\s*\d+\s*\)\s*$", "", str(a or "")).strip()
custs, stat, skipped = {}, [], []
own2cust = {}            # (section, owner) → nz(name)   · 시트 books 귀속용
owner_keys = {}          # owner → {nz(name)}            · (섹션 불명 책) 오너 유일할 때 보조 매칭
cur_key = None
for a in ALLOC:
    base = _base(a["account"])
    # named 행 = 새 고객사. SPLIT_EACH_OWNER 계정(예: Cake)은 연속행도 오너마다 개별 고객사로 쪼갠다.
    if a.get("named") or (base and nz(base) in SPLIT_EACH_OWNER):
        if not base: continue
        nm = f"{base}-{a['owner']}"; cur_key = nz(nm)
        custs.setdefault(cur_key, {"name": nm, "base": base, "sec": a["section"], "owner": a["owner"], "alloc": [], "books": []})
    if cur_key is None: continue
    custs[cur_key]["alloc"].append(a)                 # 연속행이면 위 고객사의 추가 할당
    own2cust.setdefault((a["section"], a["owner"]), []).append(cur_key)   # (s,o) 여러 고객사 claim 가능
    owner_keys.setdefault(a["owner"], set()).add(cur_key)
print(f"대장 고객사 {len(custs):,}곳 (할당 {len(ALLOC):,}행)")

def cust_of(s, o, sheet_o=None, sheet_base=None):
    """책(section,owner)을 대장 고객사에 귀속.
    ① (섹션,오너) 정확 매칭. 같은 (s,o)를 여러 고객사가 claim하면(예: humedical·엠베스트 S3/O28)
       그 책이 온 '시트명(sheet_base)'과 base가 같은 고객사를 우선 귀속.
    ② 없으면 오너-보조 매칭 — 오너가 대장에서 유일하고 '책 오너 == 시트 정본 오너'일 때만.
       (남의 시트에 잘못 적힌 이물 행 차단. 예: 네오노트(1012) 시트의 owner 12 → iconix1로 안 샘)."""
    ks = own2cust.get((s, o))
    if ks:
        if sheet_base and len(set(ks)) > 1:
            for k in ks:
                if nz(custs[k]["base"]) == nz(sheet_base): return k
        return ks[0]
    if sheet_o is not None and o != sheet_o: return None
    kk = owner_keys.get(o)
    return next(iter(kk)) if kk and len(kk) == 1 else None

# ── 수집(편집 데이터) ── 시트 books를 (section,owner)로 대장 고객사에 귀속 ──
n_orphan = 0
common_use = {}          # nz(고객사) → {"name":…, "NSP":bool, "NWP":bool}
common_books = {}        # nz(고객사) → [공유 코드에서 이 고객사가 쓰는 book 들]
for src, fname, skip in FILES:
    path = os.path.join(SRC, fname)
    if not os.path.exists(path):
        print("!! 없음:", fname); continue
    wb = load_workbook(path, data_only=True)
    n_sheet = n_book = 0
    for ws in wb.worksheets:
        if ws.title.strip().lower() in skip: continue
        cinfo = COMMON_SHEETS.get(ws.title.strip())
        name, s_hint, o_hint = sheet_customer(ws.title)
        ty = "필기펜" if src == "pen" else "소리펜"
        books = parse_sheet(ws, ty, s_hint, o_hint, common=bool(cinfo))
        if not books:
            skipped.append(f"{src}:{ws.title}"); continue
        for b in books:
            b["src"] = src
            if cinfo:
                b["cg"] = cinfo[0]                       # 커먼 코드 그룹 (NSP / NWP)
                cu = b.get("cu")
                if cu:
                    e = common_use.setdefault(nz(cu), {"name": cu, "NSP": False, "NWP": False})
                    e[cinfo[0]] = True
                    common_books.setdefault(nz(cu), []).append(b)   # 이 고객사의 공유 코드 book
            ck = cust_of(b["s"], b["o"], o_hint, name)     # 시트 정본 오너로 이물행 차단 + 시트명으로 claim 충돌 해소
            if ck: custs[ck]["books"].append(b)
            else: n_orphan += 1
        n_sheet += 1; n_book += len(books)
    stat.append((fname, n_sheet, n_book))
print(f"편집 books 대장 귀속 완료 · orphan(대장에 없는 코드) {n_orphan:,}건")

# ── 편집현황 파일 → 편집여부·심볼을 코드(k/s/o/b)로 매칭 이식 ──────────────
#   시트명(고객사)은 원장과 어긋나므로 무시하고, 코드가 일치하는 원장 책에만 적용한다.
edit_by_code = {}      # (k,s,o,b) → {"sm":[..], "pm":[..]}
epath = os.path.join(SRC, EDIT_FILE)
if os.path.exists(epath):
    ewb = load_workbook(epath, data_only=True)
    for ws in ewb.worksheets:
        for b in parse_sheet(ws, "소리펜"):
            if b.get("nb"): continue
            e = edit_by_code.setdefault((b["k"], b["s"], b["o"], b["b"]), {"sm": [0] * SOUND_N, "pm": [0] * PEN_N})
            e["sm"] = [max(a, c) for a, c in zip(e["sm"], b["sm"])]
            e["pm"] = [max(a, c) for a, c in zip(e["pm"], b["pm"])]
    # (k 불일치 대비) 코드종류를 뺀 (s,o,b) 키도 보조로 둔다 — 원장 k 를 신뢰
    edit_by_sob = {}
    for (k, s, o, bk), e in edit_by_code.items():
        t = edit_by_sob.setdefault((s, o, bk), {"sm": [0] * SOUND_N, "pm": [0] * PEN_N})
        t["sm"] = [max(a, c) for a, c in zip(t["sm"], e["sm"])]
        t["pm"] = [max(a, c) for a, c in zip(t["pm"], e["pm"])]
    n_ed = n_sym = 0
    for _c in custs.values():
        for b in _c["books"]:
            e = edit_by_code.get((b["k"], b["s"], b["o"], b["b"])) or edit_by_sob.get((b["s"], b["o"], b["b"]))
            if not e: continue
            b["ed"] = True; n_ed += 1
            if any(e["sm"]) or any(e["pm"]):
                b["sm"] = list(e["sm"]); b["pm"] = list(e["pm"]); n_sym += 1
    print(f"편집현황 코드매칭: 편집표시 {n_ed:,}건 · 심볼적용 {n_sym:,}건 (편집파일 코드 {len(edit_by_code):,}종)")
else:
    print("!! 편집현황 파일 없음:", EDIT_FILE)

# OID 는 book 으로 나누지 않는 경우가 많다(북코드 없음) → 아래 nb 필터 전에 대장용으로 따로 모아 둔다 (PC-033)
OID_ROWS = [(_c["name"], b) for _c in custs.values() for b in _c["books"] if b["k"] == "O"]

# 북코드 없는 행(중복/변형 표기, 실제 book 아님)은 코드 프로젝트·SOBP 맵·편집에서 뺀다.
#   단 **OID 는 분량이 적으면 book 을 나누지 않으므로** 그 행은 남긴다 — 편집 이력에도 나와야 한다 (PC-037)
for _c in custs.values():
    _c["books"] = [b for b in _c["books"] if (not b.get("nb")) or b["k"] == "O"]
for _k in list(common_books):
    common_books[_k] = [b for b in common_books[_k] if not b.get("nb")]
    if not common_books[_k]: del common_books[_k]

# ═══ 예외 정리: 구몬D-37 S3/O37 의 PDS3 B669 (데일리노트) — 발급된 코드가 아니다 ═══
#   담당 확인(2026-08-28): "구몬D-37 의 669 는 발급 없다" → 편집현황 표기 오류로 보고 제외한다.
#   (S/O 안에 PDS2·PDS3 가 섞여 보이던 원인 중 하나)
_n_kumon = 0
for _c in custs.values():
    before = len(_c["books"])
    _c["books"] = [b for b in _c["books"]
                   if not (b["k"] == "N" and b["s"] == 3 and b["o"] == 37 and b["b"] == 669)]
    _n_kumon += before - len(_c["books"])
if _n_kumon:
    print(f"예외 정리: 구몬D-37 S3/O37 PDS3 B669 (발급 없음) {_n_kumon}행 제외")


# (대장 기준 귀속이라 외래행 제거·오너별 분리 로직은 불필요 — 대장이 (섹션,오너)로 이미 분리됨)

# ── COMMON_SPLIT 홀더 → custs 를 코드(타입·섹션)별로 분리 ──
#   네오노트-27 → 네오노트-0-27(N/S0/O27)·네오노트-3-27(N/S3/O27)·네오노트-IDS-27(A/S4/O27)…
#   companies·projects·편집·ownership 이 모두 1회사=1코드로 일관되게 처리된다. own2cust 도 함께 갱신.
for ck in list(custs):
    c = custs[ck]
    if c["name"] not in COMMON_SPLIT: continue
    grp = {}
    for b in c["books"]:
        grp.setdefault(split_customer_name(c["name"], b["k"], b["s"]), []).append(b)
    if len(grp) <= 1: continue
    del custs[ck]
    for sn, bks in grp.items():
        nk = nz(sn)
        custs[nk] = {"name": sn, "base": c["base"], "sec": bks[0]["s"], "owner": bks[0]["o"], "alloc": [], "books": bks}
        for so in {(b["s"], b["o"]) for b in bks}:
            if own2cust.get(so): own2cust[so] = [nk if x == ck else x for x in own2cust[so]]

# ── 출력 ──
companies, projects, logs, edit_customers = [], [], [], []
cid = pid = lid = 0
T_books = T_pages = T_sym = T_size = 0

for _, c in sorted(custs.items(), key=lambda kv: -len(kv[1]["books"])):
    name, books = c["name"], c["books"]
    cid += 1
    cflag = common_use.get(nz(name), {})
    co = {"id": cid, "name": name, "manager": "", "contact": "", "address": "",
          "bizNo": "", "bankName": "", "accountNo": "", "docs": []}
    # 커먼 코드 시트에서 쓰이는 업체면 해당 체크를 켜 둔다
    if cflag.get("NSP") or any(b.get("cg") == "NSP" for b in books): co["nspCommon"] = True
    if cflag.get("NWP") or any(b.get("cg") == "NWP" for b in books): co["nwpCommon"] = True
    if nz(name) in CLOSED_ACCOUNTS:                    # 프로젝트 종료 고객사
        co["closed"] = True; co["closedNote"] = CLOSED_ACCOUNTS[nz(name)]
    companies.append(co)
    pages = sum(b["pg"] for b in books)
    size = sum(b["bytes"] for b in books)
    sym = sum(sum(b["sm"]) + sum(b["pm"]) for b in books)

    groups = {}
    for b in books:
        g = groups.setdefault((b["k"], b["s"], b["o"]), {"bmin": b["b"], "bmax": b["b"], "pages": 0, "date": b["d"]})
        g["bmin"] = min(g["bmin"], b["b"]); g["bmax"] = max(g["bmax"], b["b"]); g["pages"] += b["pg"]
        if b["d"] and (not g["date"] or b["d"] < g["date"]): g["date"] = b["d"]

    for (k, sec, own), g in sorted(groups.items()):
        pid += 1
        sub = [b for b in books if b["k"] == k and b["s"] == sec and b["o"] == own]
        _editing = any(b.get("ed") for b in sub)
        _symbols = int(sum(sum(b["sm"]) + sum(b["pm"]) for b in sub))
        # 편집 데이터(편집 프로젝트에 노출되는 코드) = casterN 편집툴 서비스
        _service = "CASTERN" if (_editing or _symbols > 0) else "NONE"
        projects.append({"id": pid, "name": f"{name} 코드발급 · S{sec}/O{own}", "companyId": cid,
                         "service": _service, "grade": "",
                         "issued": [{"id": 1, "date": g["date"] or "2020-01-01",
                                     "codes": int(g["pages"]), "used": int(g["pages"]),
                                     "kind": k, "section": sec, "owner": own,
                                     "bookStart": g["bmin"], "bookEnd": g["bmax"], "pageStart": 0, "pageEnd": 0}],
                         "editing": _editing, "editingOwner": own,
                         "symbols": _symbols})
    # 대장 할당 중 편집(책)이 없는 (섹션,오너) → 코드발급-only 프로젝트 (편집 안 됨, **코드 종류 미정**)
    #   원장에는 PDS 구분이 없다 → kind 를 넣지 않는다. 지도의 종류 배지에도 나오지 않는다 (PC-041)
    #   예) 교원도요새베트남-4: O4는 편집(위), O5는 할당만 → O5 코드발급-only.
    book_so = {(b["s"], b["o"]) for b in books}
    seen_alloc = set()
    for a in c.get("alloc", []):
        so = (a["section"], a["owner"])
        if so in book_so or so in seen_alloc: continue
        seen_alloc.add(so)
        pid += 1
        projects.append({"id": pid, "name": f"{name} 코드발급 · S{a['section']}/O{a['owner']}", "companyId": cid,
                         "service": "NONE", "grade": "",
                         "issued": [{"id": 1, "date": a.get("date") or "2020-01-01", "codes": 0, "used": 0,
                                     "section": a["section"], "owner": a["owner"],
                                     "bookStart": a.get("book_start") or 0, "bookEnd": a.get("book_end") or 0,
                                     "pageStart": 0, "pageEnd": 0}],
                         "editing": False, "editingOwner": a["owner"], "symbols": 0, "codeOnly": True})
    lid += 1
    logs.append({"id": 1000 + lid, "no": 1, "companyId": cid, "projectId": projects[-1]["id"] if projects else 0,
                 "date": "2026-03-02", "kind": "메모",
                 "content": f"{name} · 교재 {len(books):,}건 · 발급 {int(pages):,}p · 심볼 {int(sym):,}", "author": "이영업"})

    # 편집 프로젝트 고객사 1건 생성 (공통코드 홀더는 코드 타입·섹션별로 여러 건으로 분리)
    def _emit(cname, bks):
        global T_books, T_pages, T_sym, T_size
        p = sum(b["pg"] for b in bks); sz = sum(b["bytes"] for b in bks)
        sy = sum(sum(b["sm"]) + sum(b["pm"]) for b in bks)
        owners = sorted({b["o"] for b in bks}); methods = {}
        for b in bks:
            for one in [x.strip() for x in (b["m"] or "").split(",") if x.strip()]:
                methods[one] = methods.get(one, 0) + 1
        edit_customers.append({
            "customer": cname, "owner": str(owners[0] if owners else 0), "owners": owners,
            "codeKinds": sorted({b["k"] for b in bks}), "books": len(bks),
            "pages": int(p), "symbols": int(sy),
            "soundSymbols": int(sum(sum(b["sm"]) for b in bks)),
            "penSymbols": int(sum(sum(b["pm"]) for b in bks)),
            "withSymbolBooks": sum(1 for b in bks if sum(b["sm"]) + sum(b["pm"]) > 0),
            "sizeMB": round(sz / 1e6), "soundBreakdown": {}, "penBreakdown": {},
            "topMethods": sorted(methods.items(), key=lambda x: -x[1])[:4], "bookRows": bks})
        T_books += len(bks); T_pages += p; T_sym += sy; T_size += sz
    _emit(name, books)   # custs 는 위에서 코드별로 이미 분리됨

# 공유(커먼) 코드 '사용 고객사'(cu)는 별도 고객사로 등록하지 않는다.
#   고객사 명단 = 대장(오너코드_발급리스트) 뿐. cu는 공유코드 보유자 상세의 '사용 고객사'로만 표시된다.
print(f"커먼 코드 사용 고객사(cu) {len(common_use)}곳 — 고객사 미등록(공유코드 사용 표시로만 유지)")

# ── 공유(커먼) 코드 → 각 사용 고객사의 코드 프로젝트 생성 ──────────────
#   공유 코드 시트의 book 은 보유자(네오노트/PUI/NeoLAB)로 묶여 있어, cu(실사용 고객사)를
#   [코드 프로젝트]에서 선택해도 코드가 안 뜬다. cu 별로 book 을 묶어 프로젝트를 만든다.
def _runs(nums):
    xs = sorted(set(nums)); out = []
    for n in xs:
        if out and n == out[-1][1] + 1: out[-1][1] = n
        else: out.append([n, n])
    return out
id_by_name = {nz(c["name"]): c["id"] for c in companies}
have_proj = {(p["companyId"], p["issued"][0]["section"], p["issued"][0]["owner"]) for p in projects if p["issued"]}
n_common_proj = 0
for key, books in common_books.items():
    coid = id_by_name.get(key)
    if not coid: continue
    grp = {}
    for b in books:
        grp.setdefault((b["k"], b["s"], b["o"]), []).append(b)
    for (k, sec, own), bs in sorted(grp.items()):
        if (coid, sec, own) in have_proj: continue      # 이미 자체 프로젝트가 있으면 생략
        bnums = [x["b"] for x in bs]
        pmax = max((x["pg"] for x in bs), default=0)
        dates = [x["d"] for x in bs if x.get("d")]
        iss = []
        for i, (bstart, bend) in enumerate(_runs(bnums), 1):
            bc = bend - bstart + 1
            iss.append({"id": i, "date": (min(dates) if dates else "2020-01-01"),
                        "codes": bc * max(1, pmax), "used": 0, "kind": k,
                        "section": sec, "owner": own, "bookStart": bstart, "bookEnd": bend,
                        "pageStart": 0, "pageEnd": pmax})
        pid += 1
        cust_name = next((c["name"] for c in companies if c["id"] == coid), key)
        projects.append({"id": pid, "name": f"{cust_name} 공유코드 · S{sec}/O{own}",
                         "companyId": coid, "service": "NONE", "grade": "",
                         "issued": iss, "editing": False, "editingOwner": own, "symbols": 0,
                         "shared": True})
        have_proj.add((coid, sec, own)); n_common_proj += 1
print(f"공유 코드 프로젝트 {n_common_proj}건 생성")

# ── 공유 코드 실사용 고객사(cu)를 편집/코드 행에 채운다 ──────────────
#   공유 코드의 고객사명은 레퍼런스/커먼 시트의 '고객사' 열에만 있고, 편집현황 파일에는
#   보유자(NeoLAB)만 있다. (k,s,o,b) 기준으로 cu 를 편집·코드 행에 옮겨 사용 고객사가 자동 채워지게 한다.
cu_by_sob = {}
for _books in common_books.values():
    for b in _books:
        if b.get("cu"): cu_by_sob[(b["k"], b["s"], b["o"], b["b"])] = b["cu"]
n_cu_fill = 0
for c in edit_customers:
    for b in c["bookRows"]:
        if not b.get("cu"):
            v = cu_by_sob.get((b["k"], b["s"], b["o"], b["b"]))
            if v: b["cu"] = v; n_cu_fill += 1
print(f"공유 코드 사용 고객사 자동 채움 {n_cu_fill}건")

# ═══ 예외 정리: 미할당 코드 S3/O12 (901플래너 스프링) → 네오노트12 ═══════════
#   편집 서비스가 '미할당 코드'도 쓰던 시기에 실제 코드 S3/O12 로 901플래너(스프링) 편집이 진행됨.
#   편집팀 요청: 편집 실적은 네오노트-1012 에 귀속(파일이 O1012라 이미 그렇게 노출),
#   S3/O12 코드는 네오노트 소유(네오노트12)로 코드발급 처리. 향후 미할당 코드 사용 차단 정책 예정.
EXC_MEMO = ("[예외 정리 · 미할당 코드 사용 건] 편집 서비스에서 할당되지 않은 코드도 사용 가능하던 시기에, "
            "실제 코드 S3/O12 로 '901플래너(스프링)' 편집이 진행되었습니다. 편집팀 요청에 따라 편집 실적은 "
            "네오노트-1012 에 귀속(편집 프로젝트에 1012로 노출)하고, S3/O12 코드는 네오노트 소유(본 고객사 "
            "네오노트12)로 코드발급 처리합니다. 향후 미할당 코드는 편집 서비스에서 사용할 수 없도록 막을 예정입니다.")
cid += 1; _exc_cid = cid
companies.append({"id": _exc_cid, "name": "네오노트12", "manager": "", "contact": "", "address": "",
                  "bizNo": "", "bankName": "", "accountNo": "", "docs": []})
pid += 1; _exc_pid = pid
projects.append({"id": _exc_pid, "name": "네오노트12 코드발급 · S3/O12 (901플래너 스프링 · 실데이터 네오노트-1012)",
                 "companyId": _exc_cid, "service": "NONE", "grade": "",
                 "issued": [{"id": 1, "date": "2020-01-01", "codes": 0, "used": 0, "kind": "N",
                             "section": 3, "owner": 12, "bookStart": 3006, "bookEnd": 3006, "pageStart": 0, "pageEnd": 0}],
                 "editing": False, "editingOwner": 12, "symbols": 0, "codeOnly": True,
                 "editLinkOwner": 1012, "editLinkLabel": "네오노트-1012"})
lid += 1
logs.append({"id": 1000 + lid, "no": 1, "companyId": _exc_cid, "projectId": _exc_pid,
             "date": "2026-07-28", "kind": "요청", "content": EXC_MEMO, "author": "편집팀"})
# SOBP 맵(ownership-data) 에도 S3/O12 = 네오노트12 코드발급으로 등록
ALLOC.append({"account": "네오노트12", "named": True, "section": 3, "owner": 12,
              "book_start": 3006, "book_end": 3006, "page_start": 0, "page_end": 0, "date": ""})
# 편집 프로젝트의 해당 교재(901플래너 스프링, 네오노트-1012 아래)에도 업무요청 메모(히스토리) 추가
EXC_BOOK_MEMO = ("[예외 정리 · 미할당 코드] 이 교재의 실제 코드는 S3/O12 이나, 편집 서비스가 미할당 코드도 쓰던 "
                 "시기에 발생한 건으로 편집팀 요청에 따라 편집 실적은 네오노트-1012(S3/O1012)에 귀속합니다. "
                 "S3/O12 코드 자체는 고객사 '네오노트12'로 코드발급 처리했습니다. 향후 미할당 코드 사용은 차단 예정.")
_n_memo = 0
for _c in edit_customers:
    for _b in _c["bookRows"]:
        if _b.get("k") == "N" and _b.get("s") == 3 and _b.get("o") == 1012 and _b.get("b") == 3006 and "901플래너" in (_b.get("t") or ""):
            _lg = _b.setdefault("logs", [])
            _lg.append({"id": len(_lg) + 1, "no": len(_lg) + 1, "kind": "요청",
                        "content": EXC_BOOK_MEMO, "date": "2026-07-28", "author": "편집팀"})
            _n_memo += 1
print(f"예외 정리: 네오노트12 (S3/O12 코드발급) 생성 · 901플래너(스프링) 편집실적 네오노트-1012 귀속 · 교재 메모 {_n_memo}건")

json.dump({"companies": companies, "projects": projects, "logs": logs},
          open(os.path.join(WEB, "seed-customers.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)

# 편집 프로젝트 데이터 = 모든 원장(NSP+NWP) 코드 — 어느 코드든 편집·심볼 입력 대상
#   (편집현황 파일은 정본에서 제외했으므로 ed 여부로 가리지 않는다)
edit_only = []
for c in edit_customers:
    # 공통(커먼) 코드는 '프로젝트명(교재명)'이 있어야 편집 프로젝트 대상. 프로젝트명 없이 고객사(cu)만 있는
    #   행은 '코드만 확보(사용가능)' 이므로 편집 프로젝트에서 제외한다. (일반 코드는 그대로 노출)
    rows = [b for b in c["bookRows"] if not (b.get("cg") and not (b.get("t") or "").strip())]
    if not rows: continue
    pg = sum(b["pg"] for b in rows)
    snd = sum(sum(b["sm"]) for b in rows); pn = sum(sum(b["pm"]) for b in rows)
    owners = sorted({b["o"] for b in rows})
    edit_only.append({**c, "bookRows": rows, "books": len(rows), "pages": int(pg),
                      "symbols": int(snd + pn), "soundSymbols": int(snd), "penSymbols": int(pn),
                      "owner": str(owners[0] if owners else 0), "owners": owners,
                      "codeKinds": sorted({b["k"] for b in rows}),
                      "sizeMB": round(sum(b["bytes"] for b in rows) / 1e6),
                      "withSymbolBooks": sum(1 for b in rows if sum(b["sm"]) + sum(b["pm"]) > 0)})
edit_only.sort(key=lambda c: -c["symbols"])

# 번들 축소 — 빈 값/0 배열/화면에서 안 쓰는 키는 빼고 내보낸다 (앱에서 기본값 복원)
DROP_KEYS = {"src", "cg", "_do"}
def slim(b):
    o = {}
    for k, v in b.items():
        if k in DROP_KEYS: continue
        if k == "sp": o[k] = v; continue                   # 시작 페이지는 0도 유효값 → 보존
        if v in ("", None, 0, [], {}): continue
        if k in ("sm", "pm") and not any(v): continue      # 심볼 없음 → 생략
        o[k] = v
    o["b"] = b["b"]; o["s"] = b["s"]; o["o"] = b["o"]; o["k"] = b["k"]   # 키는 항상 유지
    return o
for c in edit_only:
    c["bookRows"] = [slim(b) for b in c["bookRows"]]
    c.pop("soundBreakdown", None); c.pop("penBreakdown", None)
json.dump({"summary": {"customers": len(edit_only),
                       "books": sum(c["books"] for c in edit_only),
                       "pages": sum(c["pages"] for c in edit_only),
                       "symbols": sum(c["symbols"] for c in edit_only),
                       "sizeGB": round(sum(c["sizeMB"] for c in edit_only) / 1000, 1)},
           "customers": edit_only},
          open(os.path.join(WEB, "editing-detail.json"), "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))
print("editing-detail.json", round(os.path.getsize(os.path.join(WEB, "editing-detail.json")) / 1024), "KB ·", len(edit_only), "고객사")

# 지도/목록용 경량 데이터 (번들 최소화): k,s,o,b,p(page),sp,c(보유업체),t(교재),e(편집),cu(공유코드 실사용 고객사),pen(소리펜/필기펜)
#   k = 좌표(SOBP)의 코드 종류: N=PDS3 · G=PDS2 · A=IDS(이력) · O=OID(인덱스 전용). PDS4(S-code)는 Section 44 로 판별.
def crow(c, b):
    if b.get("nb"): return None                        # 북코드 없는 행은 지도(SOBP 맵)에 표시하지 않음
    r = {"k": b["k"], "s": b["s"], "o": b["o"], "b": b["b"], "p": b["pg"],
         "sp": b.get("sp", 0), "c": c["customer"], "t": (b["t"] or "")[:40],
         "e": 1 if b.get("ed") else 0}
    # 펜 구분(좌표 속성) — S=소리펜(NSP) · W=필기펜(NWP). 원장 파일(src)이 정본.
    if b.get("src"): r["pen"] = "W" if b["src"] == "pen" else "S"
    if b.get("cu"): r["cu"] = b["cu"][:40]        # 공유(커먼) 코드에서 실제 사용 고객사
    if "ea" in b: r["ea"] = b["ea"]               # 레퍼런스 시트 편집 O/X (1=편집, 0=코드만 할당)
    if not r["sp"]: del r["sp"]
    return r
#   OID(k="O")도 좌표 종류의 하나로 지도·목록에서 필터한다 (PC-035). 업체별 index 목록은 oid-data.json.
compact = [r for c in edit_customers for b in c["bookRows"] if (r := crow(c, b))]
# OID 는 분량이 적으면 book 을 나누지 않는다(북코드 없음) → 그런 행도 지도에 싣되 nb=1 로 표시한다.
#   (book 카드가 아니라 OWNER 단위로만 노출 — 예: 한솔교육 S3/O25)
n_oid_nb = 0
for _name, b in OID_ROWS:
    if not b.get("nb"):
        continue                                   # 북코드가 있는 행은 위 compact 에 이미 있다
    r = {"k": "O", "s": b["s"], "o": b["o"], "b": 0, "p": b.get("pg", 0),
         "c": _name, "t": (b["t"] or "")[:40], "e": 1 if b.get("ed") else 0, "nb": 1}
    if b.get("src"): r["pen"] = "W" if b["src"] == "pen" else "S"
    if b.get("cu"): r["cu"] = b["cu"][:40]
    compact.append(r); n_oid_nb += 1

json.dump(compact, open(os.path.join(WEB, "code-usage.json"), "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))
print("code-usage.json", round(os.path.getsize(os.path.join(WEB, "code-usage.json")) / 1024), "KB ·", len(compact), "행",
      f"(OID book 미분할 {n_oid_nb}행 포함)")

# ── OID 관리대장 → oid-data.json ────────────────────────────────────
#   OID = index 만 갖는 코드(외부 코드를 우리 펜으로 읽기 위한 방식). 우리 펜으로 찍으면 코드 값이 1개만 나온다.
#   총량이 약 6만 개뿐이라 책이 많지 않으면 book 으로 나누지 않는다.
#   대장 관리 방식: 업체 구분은 S/O 로 해 왔고, 분량이 늘어난 업체만 book 번호(=OID index)로 나눈다.
#   → SOBP(PDS2·PDS3·PDS4)와 같은 좌표 관리가 아니라 **업체 + index** 로 관리하고, 기존 좌표는 메모로 남긴다.
OID_TOTAL = 60000        # OID 코드 총량(약 6만) — 사용량 대비 표시용
oid_group = {}
for _name, b in OID_ROWS:
    if True:
        key = _name
        g = oid_group.setdefault(key, {"company": key, "section": b["s"], "owner": b["o"],
                                       "pen": "W" if b.get("src") == "pen" else "S", "items": []})
        item = {"title": (b["t"] or "")[:60]}
        if not b.get("nb"):
            item["idx"] = b["b"]                       # book 번호 = OID index 로 관리하는 업체
        if b.get("pg"): item["pages"] = b["pg"]
        if b.get("d"): item["date"] = b["d"][:10]
        if b.get("cu"): item["cu"] = b["cu"][:40]
        g["items"].append(item)
for g in oid_group.values():
    idxs = [i["idx"] for i in g["items"] if "idx" in i]
    g["indexBy"] = "book" if idxs else "none"          # book 번호로 index 관리 / 미분할(업체 단위)
    g["indexRange"] = [min(idxs), max(idxs)] if idxs else None
    g["count"] = len(g["items"])
    # 기존 SOBP 히스토리는 메모로만 남긴다 (좌표 관리 대상 아님)
    g["sobpMemo"] = ("S%s/O%s" % (g["section"], g["owner"])) + (
        " · B%s~%s" % (g["indexRange"][0], g["indexRange"][1]) if g["indexRange"] else " · book 미분할")
oid_json = {"total": OID_TOTAL,
            "companies": sorted(oid_group.values(), key=lambda x: -x["count"]),
            "meta": {"source": "NSP/NWP 원장의 코드 구분 = OID 행",
                     "note": "OID 는 index 전용 코드다. SOBP(PDS2·PDS3·PDS4) 좌표 관리와 분리해 업체+index 로 관리한다(PC-033). 기존 좌표는 메모."}}
json.dump(oid_json, open(os.path.join(WEB, "oid-data.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("oid-data.json · 업체", len(oid_json["companies"]), "· 항목",
      sum(g["count"] for g in oid_json["companies"]))

# ── 공통(커먼) 코드 레지스트리 + 사용 고객사 멤버십(히스토리→시드) → common-members.json ──
#   코드 정본 = (타입 k: N=PDS3·G=PDS2·A=IDS, section, owner). 사용 고객사(cu)는 엑셀 히스토리에서 자동 시드.
#   web/lib/commonCodes.ts 의 COMMON_CODES 와 (k,s,o) 가 일치해야 한다.
COMMON_CODES = [("G", 3, 21), ("N", 0, 27), ("N", 3, 27), ("N", 3, 1012), ("N", 14, 303), ("A", 4, 27),
                ("N", 3, 1013), ("N", 3, 940), ("G", 3, 37), ("G", 3, 964)]
_mem = {f"{k}:{s}:{o}": set() for (k, s, o) in COMMON_CODES}
for r in compact:
    ck = f"{r['k']}:{r['s']}:{r['o']}"
    if ck in _mem and r.get("cu"): _mem[ck].add(r["cu"])
member_json = {ck: sorted(v) for ck, v in _mem.items()}
json.dump(member_json, open(os.path.join(WEB, "common-members.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("common-members.json ·", " · ".join(f"{ck}={len(v)}" for ck, v in member_json.items()))

# ── 하위(공통코드 사용) 고객사도 정식 고객사(회사 레코드)로 생성 ──────────────
#   하위 고객사도 일반 고객사처럼 상세 정보(담당자·사업자·주소)를 갖는다. 코드/SOBP는 코드 프로젝트·SOBP 맵에서 관리.
#   회사 레코드가 없던 히스토리 멤버(cu)에 대해 최소 정보의 회사를 만들어 이름순 목록에 함께 노출·검색·상세 가능하게 한다.
_next_cid = max((c["id"] for c in companies), default=0)
_existing_nz = {nz(c["name"]) for c in companies}
_sub_added = 0
for _ck, _names in member_json.items():
    for _cu in _names:
        if not _cu or nz(_cu) in _existing_nz: continue    # 이미 회사로 있으면(대장/일반) 그 회사로 노출
        _existing_nz.add(nz(_cu))
        _next_cid += 1
        companies.append({"id": _next_cid, "name": _cu, "manager": "", "contact": "", "address": "",
                          "bizNo": "", "bankName": "", "accountNo": "", "docs": [], "sub": True})
        _sub_added += 1
print(f"하위(공통코드 사용) 고객사 회사레코드 생성 {_sub_added}건 → 총 고객사 {len(companies)}")

for f, ns, nb in stat: print(f"  {f}: 시트 {ns} · 교재 {nb:,}")
if skipped: print("  (미인식/빈 시트):", ", ".join(skipped))
print(f"고객사 {len(companies)} · 프로젝트 {len(projects)} · 교재 {T_books:,} · 페이지 {int(T_pages):,} · 심볼 {int(T_sym):,}")

# ── 할당 원장(오너코드_발급리스트) → ownership-data.json ──
#   SOBP 맵 owner 표기 통일: account 를 모델 고객사명({base}-{owner})으로 (code-usage 와 동일).
def _acct_name(a):
    # own2cust 는 (s,o)→[여러 claim] 이므로 cust_of 로 단일 고객사 확정
    # (같은 S/O를 여러 곳이 claim하면 이 행이 온 account(sheet_base)와 base가 같은 곳 우선)
    ck = cust_of(a["section"], a["owner"], sheet_base=_base(a["account"]))
    return custs[ck]["name"] if ck else f'{_base(a["account"])}-{a["owner"]}'
secmap = {}
for a in ALLOC:
    secmap.setdefault(a["section"], []).append({
        "account": _acct_name(a), "owner": a["owner"], "product": "UNKNOWN",
        "book_start": a["book_start"], "book_end": a["book_end"],
        "page_start": a["page_start"], "page_end": a["page_end"],
        "shared": False})   # 예약(RESERVED) 개념 폐기 — 원장의 코드는 모두 '할당됨'(PC-004)
# 섹션별 owner 정원(코드 관리 정보) — 없으면 관측 최대치
OWNER_CAP = {0: 1024, 1: 1024, 3: 4096, 5: 256, 10: 1024, 11: 1024, 14: 4096, 15: 32768, 44: 1024}
acc_agg = {}
for a in ALLOC:
    x = acc_agg.setdefault(a["account"], {"name": a["account"], "owners": set(), "books": 0})
    x["owners"].add((a["section"], a["owner"]))
    if a["book_start"] is not None:
        x["books"] += max(0, (a["book_end"] or 0) - (a["book_start"] or 0) + 1)
accounts = sorted(({"name": v["name"], "owners": len(v["owners"]), "books": v["books"]} for v in acc_agg.values()),
                  key=lambda x: -x["books"])
own_json = {"sections": [{"section": s2, "legacy": False, "test_dev": s2 in (1, 44),
                          "axis_max": max(OWNER_CAP.get(s2, max((r["owner"] for r in recs), default=0) + 1) - 1, 0),
                          "total_owners": OWNER_CAP.get(s2, max((r["owner"] for r in recs), default=0) + 1),
                          "owned": len({r["owner"] for r in recs}), "over_dim": False,
                          "records": recs}
                         for s2, recs in sorted(secmap.items())],
            "accounts": accounts,
            "meta": {"source": "오너코드_발급리스트", "rows": len(ALLOC),
                     "records": len(ALLOC), "account_count": len(accounts),
                     "code_state_note": "원장의 모든 코드 = 할당됨(발급). 예약/사용중 구분 없음(PC-004)."}}
json.dump(own_json, open(os.path.join(WEB, "ownership-data.json"), "w", encoding="utf-8"), ensure_ascii=False)

# 할당 원장에만 있는 (고객사, S, O) → 코드 프로젝트에 보강
by_name = {nz(c["name"]): c for c in companies}
have = {(p["companyId"], p["issued"][0]["section"], p["issued"][0]["owner"]) for p in projects}
added = 0
for a in ALLOC:
    co = by_name.get(nz(a["account"]))
    if not co: continue
    key = (co["id"], a["section"], a["owner"])
    if key in have:
        # 이미 있는 프로젝트는 할당 원장 범위로 보정 (원장이 실제 할당 기준)
        for p2 in projects:
            if p2["companyId"] == co["id"] and p2["issued"][0]["section"] == a["section"] and p2["issued"][0]["owner"] == a["owner"]:
                iss = p2["issued"][0]
                if a["book_start"] is not None: iss["bookStart"] = a["book_start"]; iss["bookEnd"] = a["book_end"]
                if a["page_start"] is not None: iss["pageStart"] = a["page_start"]; iss["pageEnd"] = a["page_end"]
                bcnt = max(0, (iss["bookEnd"] or 0) - (iss["bookStart"] or 0) + 1)
                pcnt = max(0, (iss["pageEnd"] or 0) - (iss["pageStart"] or 0) + 1)
                if bcnt and pcnt: iss["codes"] = bcnt * pcnt      # 발급 규모 = B × P
        continue
    pid += 1; have.add(key)
    # 편집 데이터로 종류를 확정할 수 있을 때만 넣는다 — 원장에는 PDS 구분이 없다 (PC-041)
    kinds = [b["k"] for c in edit_customers if nz(c["customer"]) == nz(a["account"]) for b in c["bookRows"]]
    k = max(set(kinds), key=kinds.count) if kinds else None
    bs, be = a["book_start"] or 0, a["book_end"] or 0
    ps, pe = a["page_start"] or 0, a["page_end"] or 0
    projects.append({"id": pid, "name": f"{co['name']} 코드발급 · S{a['section']}/O{a['owner']}",
                     "companyId": co["id"], "service": "NONE", "grade": "",
                     # 발급 코드 규모 = Book수 × Page수 (할당 원장 기준)
                     "issued": [{"id": 1, "date": a["date"] or "2020-01-01",
                                 "codes": max(0, be - bs + 1) * max(0, pe - ps + 1), "used": 0,
                                 **({"kind": k} if k else {}), "section": a["section"], "owner": a["owner"],
                                 "bookStart": bs, "bookEnd": be, "pageStart": ps, "pageEnd": pe}],
                     "editing": False, "editingOwner": a["owner"], "symbols": 0})
    added += 1
json.dump({"companies": companies, "projects": projects, "logs": logs},
          open(os.path.join(WEB, "seed-customers.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"할당 원장 {len(ALLOC)}행 → ownership-data.json · 프로젝트 보강 {added}건")

# ── 언어 슬롯 매핑 → language-slots.json ──
lang_json = {"pds": LANG_SLOTS_PDS, "section": LANG_SLOTS_SECTION, "baseOwner": LANG_BASE_OWNER,
             "slots": [{"owner": o, "slot": sl, "lang": lg, "base": o == LANG_BASE_OWNER}
                       for (o, sl, lg) in sorted(LANG_SLOTS, key=lambda x: x[1])]}
json.dump(lang_json, open(os.path.join(WEB, "language-slots.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
_owners = sorted({o for (o, _, _) in LANG_SLOTS})
print(f"언어 슬롯 {len(LANG_SLOTS)}개 (owner {_owners}) → language-slots.json")
