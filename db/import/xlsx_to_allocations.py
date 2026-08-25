#!/usr/bin/env python3
# NcodeCenter — 오너코드 발급 리스트 XLSX(다중 시트) → allocations 변환/검증
# 사용: python db/import/xlsx_to_allocations.py "<xlsx경로>" [--sql]
#   - 각 시트를 순회. 데이터행 = 2번째 칸이 PDS2/PDS3 인 행(헤더 줄 수 무관).
#   - 컬럼(위치): 0 고객사,1 코드구분(product),2 Section,3 Owner,4 Book,
#                 5 Start Page,6 Total Page,7 발급일자,8 삭제일자,9 교재명 ...
#   - Book/Page는 1-based. page_end = Start + Total - 1.
import sys, re
from openpyxl import load_workbook

# Windows 콘솔(cp949)에서도 한글/기호 출력되도록 UTF-8 고정
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

DIM = {
  "PDS3":{0:(1023,16383,4095),3:(1023,8191,511),5:(255,4095,4095),10:(1023,4095,1023),
          11:(1023,8191,511),14:(1023,8191,31),15:(32767,4095,511)},
  "PDS2":{0:(524287,8191,1023),3:(4095,4095,4095),14:(4095,4095,1023)},
}  # (owner_max, book_max, page_max)

def s(v):
    if v is None: return ""
    if hasattr(v, "date"): return str(v.date())      # datetime → YYYY-MM-DD
    return str(v).strip()

def num(v):
    try: return int(float(v))
    except: return None

def main():
    if len(sys.argv) < 2:
        print('사용: python xlsx_to_allocations.py "<xlsx>" [--sql]'); sys.exit(2)
    path = sys.argv[1]; emit_sql = "--sql" in sys.argv[2:]
    wb = load_workbook(path, data_only=True, read_only=True)

    allocs, errors, skipped = [], [], []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        data = [r for r in rows if len(r) > 1 and r[1] and re.match(r"^PDS\d$", str(r[1]).strip(), re.I)]
        if not data:
            skipped.append(ws.title); continue
        for r in data:
            prod = str(r[1]).strip().upper()
            sec, own, book = num(r[2]), num(r[3]), num(r[4])
            sp, tp = num(r[5]), num(r[6])
            if None in (sec, own, book, sp, tp):
                errors.append(f"[{ws.title}] 숫자 파싱 실패: {r[:7]}"); continue
            a = dict(sheet=ws.title, customer=s(r[0]), product=prod, section=sec, owner=own,
                     book_start=book, book_end=book, page_start=sp, page_end=sp+tp-1, total_page=tp,
                     issued_at=s(r[7]) or None, deleted_at=s(r[8]) or None, book_name=s(r[9]),
                     status="DELETED" if s(r[8]) else "ACTIVE")
            d = DIM.get(prod, {}).get(sec)
            if not d: errors.append(f"[{ws.title}] 알수없는 {prod}/S{sec}")
            else:
                om, bm, pm = d
                if own > om: errors.append(f"[{ws.title}] Owner 초과 {own}>{om}")
                if a["book_end"] > bm: errors.append(f"[{ws.title}] Book 초과 {a['book_end']}>{bm}")
                if a["page_end"] > pm: errors.append(f"[{ws.title}] Page 초과 {a['page_end']}>{pm} (B{book})")
            allocs.append(a)

    # 요약
    print(f"=== 시트: 데이터 {len(wb.worksheets)-len(skipped)} / 전체 {len(wb.worksheets)} ===")
    if skipped: print(f"  (데이터 없음 스킵: {', '.join(skipped)})")
    groups = {}
    for a in allocs:
        k = (a["customer"], a["product"], a["section"], a["owner"])
        g = groups.setdefault(k, {"books":0,"pages":0})
        g["books"] += 1; g["pages"] += a["total_page"]
    print(f"\n=== 묶음 (고객/제품/Section/Owner) : {len(groups)}개 ===")
    for (cust,prod,sec,own),g in sorted(groups.items()):
        print(f"  {cust:16} {prod} S{sec}/O{own} · {g['books']} book · {g['pages']} page")
    print(f"\n총 allocations 행: {len(allocs)} · 고객 {len({a['customer'] for a in allocs})} · "
          f"총 페이지 {sum(a['total_page'] for a in allocs)}")
    print(f"상태: {sum(a['status']=='ACTIVE' for a in allocs)} ACTIVE / {sum(a['status']=='DELETED' for a in allocs)} DELETED")
    print("\n=== 검증 ===")
    print("  ✓ 범위/무결성 오류 없음" if not errors else "\n".join("  ⚠ "+e for e in errors[:50]))

    if emit_sql:
        print("\n=== SQL ===")
        for c in sorted({a["customer"] for a in allocs}):
            print(f"INSERT INTO customers(name,service_type,status) VALUES ('{c}','FORMSOLUTION','ACTIVE') ON CONFLICT DO NOTHING;")
        for a in allocs:
            print(f"INSERT INTO allocations(customer_id,product,section,owner,book_start,book_end,page_start,page_end,source,status) "
                  f"SELECT id,'{a['product']}',{a['section']},{a['owner']},{a['book_start']},{a['book_end']},"
                  f"{a['page_start']},{a['page_end']},'IMPORT','{a['status']}' FROM customers WHERE name='{a['customer']}';")

if __name__ == "__main__":
    main()
