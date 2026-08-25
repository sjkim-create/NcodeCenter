#!/usr/bin/env python3
# NcodeCenter — 소유권 데이터(JSON) 생성: 마스터(레거시) + 상세시트(활성, 제품 확인)
# 사용: python db/import/build_ownership_data.py "<xlsx>"
# 출력(데이터만): web/data/ownership-data.json
#   - 마스터(오너코드_발급리스트): ACCOUNT×Section×Owner×book/page범위 (제품 컬럼 없음→UNKNOWN)
#   - 상세시트(S{sec}_O{own}_{고객}): 시트명에서 고객/Section/Owner, 셀에서 제품(PDS2/3) 추출
#   - (Section,Owner)가 상세시트에 있으면 상세(제품 확인) 우선, 없으면 마스터(레거시).
import sys, re, json, os
from collections import defaultdict
from openpyxl import load_workbook
try: sys.stdout.reconfigure(encoding="utf-8")
except Exception: pass

MASTER="오너코드_발급리스트"
NON_OWNER={MASTER,"티켓발급리스트","P2UI Param"}
DIM_PDS3={0:1023,3:1023,5:255,10:1023,11:1023,14:1023,15:32767}
KNOWN=set(DIM_PDS3)
# 상용 미출시 테스트/개발 전용 Section (정보표에 없음, 개발만 되고 서비스 미출시)
TEST_DEV={1,44}

def num(v):
    try: return int(float(v))
    except: return None
def rng(s):
    if s is None: return [None,None]
    m=re.search(r"(\d[\d,]*)\s*~\s*(\d[\d,]*)", str(s))
    if m: return [int(m.group(1).replace(",","")), int(m.group(2).replace(",",""))]
    m2=re.fullmatch(r"\s*(\d[\d,]*)\s*", str(s))
    if m2: v=int(m2.group(1).replace(",","")); return [v,v]
    return [None,None]
def parse_name(name):
    secs=[]; owns=[]; cust=None
    m=re.match(r"S([\d,]+)_O([\d~]+)_?(.*)", name)
    if m:
        secs=[int(x) for x in m.group(1).split(",") if x!=""]
        o=m.group(2)
        owns=list(range(int(o.split("~")[0]),int(o.split("~")[1])+1)) if "~" in o else [int(o)]
        cust=(m.group(3) or "").strip() or None
    return secs, owns, cust

def main():
    xlsx=sys.argv[1]
    wb=load_workbook(xlsx, data_only=True, read_only=True)

    # 상세시트 → (section,owner): [ (고객, 제품, book수 근사) ]  (충돌 시 둘 다 보존)
    detail=defaultdict(list)   # (s,o) -> list of dict(customer, product, books, sheet)
    for ws in wb.worksheets:
        if ws.title in NON_OWNER: continue
        secs,owns,cust=parse_name(ws.title)
        if not secs or not owns: continue
        prods=set(); bookcount=0
        for r in ws.iter_rows(values_only=True):
            has_pds=False
            for v in list(r)[:6]:
                if v is not None and re.fullmatch(r"PDS\d",str(v).strip(),re.I):
                    prods.add(str(v).strip().upper()); has_pds=True
            if has_pds: bookcount+=1
        product = "MIXED" if len(prods)>1 else (next(iter(prods)) if prods else "UNKNOWN")
        for s in secs:
            for o in owns:
                detail[(s,o)].append(dict(customer=cust or ws.title, product=product,
                                          books=max(1,round(bookcount/max(1,len(owns)))), sheet=ws.title))

    # 마스터 → (section,owner) 레코드
    ws=wb[MASTER]; rows=list(ws.iter_rows(values_only=True))
    master=[]
    for r in rows[3:]:
        if not r or r[1] in (None,""): continue
        sec=num(r[2]); own=num(r[3])
        if sec is None or own is None: continue
        b=rng(r[4])
        master.append(dict(account=str(r[1]).strip(), section=sec, owner=own,
                           book_start=b[0], book_end=b[1]))

    # 병합: 상세시트(제품 확인) 우선, 없으면 마스터(레거시). 모두 '할당됨'(예약 개념 폐기, PC-004).
    #  같은 (s,o)에 상세 여러 고객이면 book 단위로 나눠 쓰는 '공유 owner'(충돌 아님) — 모두 보존.
    recs=[]; seen=set(); shared=[]
    for (s,o),lst in detail.items():
        custs=sorted({d["customer"] for d in lst})
        is_shared = len(custs)>1
        if is_shared: shared.append((s,o,custs))
        for d in lst:
            recs.append(dict(account=d["customer"], section=s, owner=o, product=d["product"],
                             book_start=None, book_end=None, books=d["books"],
                             source="detail", shared=is_shared))
        seen.add((s,o))
    for m in master:
        if (m["section"],m["owner"]) in seen: continue
        bk = (m["book_end"]-m["book_start"]+1) if (m["book_start"] is not None) else 0
        recs.append(dict(account=m["account"], section=m["section"], owner=m["owner"],
                         product="UNKNOWN", book_start=m["book_start"], book_end=m["book_end"],
                         books=bk, source="master", shared=False))

    # 섹션 메타
    bySec=defaultdict(list)
    for x in recs: bySec[x["section"]].append(x)
    sections=[]
    for sec in sorted(bySec):
        owners=[x["owner"] for x in bySec[sec]]; dmax=max(owners); distinct=sorted(set(owners))
        known=sec in KNOWN
        sections.append(dict(section=sec, legacy=(not known), test_dev=(sec in TEST_DEV),
            axis_max=(max(dmax,DIM_PDS3[sec]) if known else dmax),
            total_owners=(DIM_PDS3[sec]+1) if known else None,
            owned=len(distinct), over_dim=(known and dmax>DIM_PDS3[sec]),
            records=[dict(account=x["account"],owner=x["owner"],product=x["product"],
                          book_start=x["book_start"],book_end=x["book_end"],
                          shared=x.get("shared",False)) for x in bySec[sec]]))
    # ACCOUNT 점유
    ao=defaultdict(set); ab=defaultdict(int)
    for x in recs:
        ao[x["account"]].add((x["section"],x["owner"])); ab[x["account"]]+=x.get("books",0)
    accounts=[dict(name=a,owners=len(ao[a]),books=ab[a]) for a in sorted(ao,key=lambda a:(-len(ao[a]),a.lower()))]

    prodcount=defaultdict(int)
    for x in recs: prodcount[x["product"]]+=1

    data=dict(sections=sections, accounts=accounts,
              meta=dict(records=len(recs), account_count=len(accounts),
                        section_ids=[s["section"] for s in sections],
                        product_breakdown=dict(prodcount),
                        shared_owners=[dict(section=s,owner=o,accounts=c) for s,o,c in shared],
                        code_state_note="모든 코드 = 할당됨(발급). 예약/사용중 구분 없음(PC-004).",
                        product_source="상세시트 우선(제품 확인)·나머지 마스터(UNKNOWN)"))

    root=os.path.abspath(os.path.join(os.path.dirname(__file__),"..",".."))
    outdir=os.path.join(root,"web","data"); os.makedirs(outdir,exist_ok=True)
    out=os.path.join(outdir,"ownership-data.json")
    with open(out,"w",encoding="utf-8") as f: json.dump(data,f,ensure_ascii=False,indent=1)
    print(f"레코드 {len(recs)} · ACCOUNT {len(accounts)} · Section {data['meta']['section_ids']}")
    print(f"코드 상태: 모두 할당됨(발급) — 예약 개념 없음")
    print(f"제품: {dict(prodcount)}")
    if shared:
        print(f"공유 owner {len(shared)}건 (여러 프로젝트가 book 단위로 나눠 씀):")
        for s,o,c in shared: print(f"    S{s}/O{o}: {c}")
    print(f"→ {out}")

if __name__=="__main__":
    main()
