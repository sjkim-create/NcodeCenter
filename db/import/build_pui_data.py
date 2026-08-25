# PUI_Ncode_List.xlsx → 피지컬 조작(PUI) 코드 할당 데이터
# 시트 = PDS{n}_S{section}_O{owner}[_라벨] · 프로젝트표 + 기능표
import sys, os, re, json
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.join(HERE, "..", "..", "web", "data")
NAME = "PUI_Ncode_List.xlsx"
_c = [os.path.join(HERE, "..", "source", NAME), os.path.join(r"C:\Users\NeoLab\Downloads", NAME)]
SRC = next((p for p in _c if os.path.exists(p)), _c[0])
print("원본:", os.path.abspath(SRC))

def s(v):
    return "" if v is None else str(v).replace("\n", " ").strip()
def n(v):
    try:
        if v in (None, "", "-"): return None
        return int(float(v))
    except Exception: return None

def find_col(hdr, *keys):
    for i, h in enumerate(hdr):
        t = s(h).replace(" ", "").lower()
        for k in keys:
            if k.replace(" ", "").lower() in t: return i
    return None

wb = load_workbook(SRC, data_only=True)
allocs = []
for name in wb.sheetnames:
    m = re.match(r"PDS(\d)_S(\d+)_O(\d+)(?:_(.*))?$", name)
    if not m: continue
    if "사본" in name: continue
    pds, sec, owner, label = m.group(1), int(m.group(2)), int(m.group(3)), (m.group(4) or "").strip()
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 600), values_only=True))
    pi = fi = None
    for i, r in enumerate(rows):
        cells = [s(x) for x in r[:12]]
        # 헤더 행만 인식: 셀이 정확히 PROJECT 이거나 '프로젝트'를 포함
        if pi is None and any(c.strip().upper() == "PROJECT" or "프로젝트" in c for c in cells): pi = i
        if fi is None and any("대구분" in c for c in cells): fi = i
    # 원본 표 보존 (시트마다 형식이 달라 그대로 보여주기 위함)
    raw = []
    for r in rows[:300]:
        cells = [s(x) for x in r[:14]]
        if any(cells): raw.append(cells)

    projects, funcs = [], []
    # 프로젝트 헤더 재탐색 (PROJECT / 프로젝트 / 프로젝트명 / 프로젝트 / 패키지)
    if pi is None:
        for i, r in enumerate(rows):
            j = " ".join(s(x) for x in r[:12])
            if "프로젝트" in j: pi = i; break
    if pi is not None:
        ph = list(rows[pi])
        cP = find_col(ph, "PROJECT", "프로젝트")
        cPage, cProd = find_col(ph, "PAGE", "페이지"), find_col(ph, "PRODUCT", "제품")
        cBook = find_col(ph, "book", "북코드")
        cMemo = find_col(ph, "메모", "포함기능")
        cDept, cNote = find_col(ph, "발급요청부서", "부서", "발급인"), find_col(ph, "비고", "데모용")
        cCust = find_col(ph, "고객사")
        end = fi if fi is not None and fi > pi else len(rows)
        for r in rows[pi + 1:end]:
            nm = s(r[cP]) if cP is not None and cP < len(r) else ""
            if not nm or nm.startswith("프로젝트"): continue
            projects.append({
                "project": nm,
                "page": s(r[cPage]) if cPage is not None and cPage < len(r) else "",
                "book": s(r[cBook]) if cBook is not None and cBook < len(r) else "",
                "product": s(r[cProd]) if cProd is not None and cProd < len(r) else "",
                "customer": s(r[cCust]) if cCust is not None and cCust < len(r) else "",
                "memo": s(r[cMemo]) if cMemo is not None and cMemo < len(r) else "",
                "dept": s(r[cDept]) if cDept is not None and cDept < len(r) else "",
                "note": s(r[cNote]) if cNote is not None and cNote < len(r) else "",
            })
        # 중복 프로젝트명 정리(같은 이름 연속 반복 제거)
        seen = set(); ded = []
        for p in projects:
            key = (p["project"], p["page"], p["book"])
            if key in seen: continue
            seen.add(key); ded.append(p)
        projects = ded[:200]
    if fi is not None:
        fh = list(rows[fi])
        c1, c2 = find_col(fh, "대구분"), find_col(fh, "소구분")
        cName, cSum, cDesc = find_col(fh, "기능명칭", "기능 명칭"), find_col(fh, "기능요약"), find_col(fh, "추가설명")
        cBook, cPg, cPar = find_col(fh, "Book"), find_col(fh, "Page"), find_col(fh, "params")
        cat = sub = ""
        for r in rows[fi + 1:]:
            nm = s(r[cName]) if cName is not None and cName < len(r) else ""
            if not nm: continue
            if c1 is not None and c1 < len(r) and s(r[c1]): cat = s(r[c1])
            if c2 is not None and c2 < len(r) and s(r[c2]): sub = s(r[c2])
            funcs.append({
                "cat": cat, "sub": sub, "name": nm,
                "summary": s(r[cSum]) if cSum is not None and cSum < len(r) else "",
                "desc": s(r[cDesc]) if cDesc is not None and cDesc < len(r) else "",
                "book": n(r[cBook]) if cBook is not None and cBook < len(r) else None,
                "page": n(r[cPg]) if cPg is not None and cPg < len(r) else None,
                "params": s(r[cPar]) if cPar is not None and cPar < len(r) else "",
            })
    allocs.append({"sheet": name, "pds": f"PDS{pds}", "section": sec, "owner": owner, "label": label,
                   "projects": projects, "funcs": funcs, "raw": raw, "images": len(getattr(ws, "_images", []))})

allocs.sort(key=lambda a: (a["section"], a["owner"]))
out = {"allocations": allocs, "summary": {"sheets": len(allocs),
       "projects": sum(len(a["projects"]) for a in allocs), "funcs": sum(len(a["funcs"]) for a in allocs)}}
json.dump(out, open(os.path.join(WEB, "pui-data.json"), "w", encoding="utf-8"), ensure_ascii=False)
print(f"할당 {len(allocs)}건 · 프로젝트 {out['summary']['projects']} · 기능 {out['summary']['funcs']}")
for a in allocs: print(f"  {a['sheet']}: S{a['section']}/O{a['owner']} {a['label']} · 프로젝트 {len(a['projects'])} · 기능 {len(a['funcs'])}")
