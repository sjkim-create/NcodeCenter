# 2_New_NSP_Ncode_List.xlsx → 편집 프로젝트(심볼 기반 작업량) 집계
# 시트=고객사, 행=책(북코드), 심볼 갯수(col14~24)=편집 작업량
import sys, os, re, json
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

SRC = r"C:\Users\NeoLab\Downloads\2_New_NSP_Ncode_List.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "web", "data", "editing-data.json")

SOUND = [("기본", 13), ("멀티터치·번역", 14), ("슬롯전환", 15), ("전체듣기", 16), ("게임", 17), ("프롬프트", 18), ("발음평가", 19)]
PEN = [("필기펜 기본", 20), ("캘린더", 21), ("링크", 22), ("교원구몬/KEP", 23)]

def num(v):
    try:
        if v in (None, "", "-"): return 0.0
        return float(v)
    except Exception:
        return 0.0

wb = load_workbook(SRC, data_only=True)
customers = []
G_sound = {k: 0 for k, _ in SOUND}
G_pen = {k: 0 for k, _ in PEN}
T_books = T_pages = T_sym = T_size = 0

for ws in wb.worksheets:
    m = re.search(r"\((\d+)\)", ws.title)
    owner = m.group(1) if m else ""
    name = re.sub(r"\(\d+\)", "", ws.title).strip()
    books = pages = sym = size = withsym = 0
    smode = {k: 0 for k, _ in SOUND}
    pmode = {k: 0 for k, _ in PEN}
    kinds = set(); methods = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or r[0] in (None, ""): continue
        books += 1
        pages += num(r[6]); size += num(r[12])
        s = num(r[24]); sym += s
        if s > 0: withsym += 1
        if r[1]: kinds.add("N" if "PDS3" in str(r[1]) else "G")
        for k, i in SOUND:
            v = num(r[i]) if len(r) > i else 0; smode[k] += v; G_sound[k] += v
        for k, i in PEN:
            v = num(r[i]) if len(r) > i else 0; pmode[k] += v; G_pen[k] += v
        if len(r) > 25 and r[25]:
            key = str(r[25]).split(",")[0].strip()
            methods[key] = methods.get(key, 0) + 1
    sound_t = int(sum(smode.values())); pen_t = int(sum(pmode.values()))
    customers.append({
        "customer": name, "owner": owner,
        "codeKinds": sorted(kinds), "books": books, "pages": int(pages),
        "symbols": int(sym), "soundSymbols": sound_t, "penSymbols": pen_t,
        "withSymbolBooks": withsym, "sizeMB": round(size / 1e6),
        "soundBreakdown": {k: int(v) for k, v in smode.items()},
        "penBreakdown": {k: int(v) for k, v in pmode.items()},
        "topMethods": sorted(methods.items(), key=lambda x: -x[1])[:4],
    })
    T_books += books; T_pages += pages; T_sym += sym; T_size += size

customers.sort(key=lambda c: -c["symbols"])
out = {
    "summary": {
        "customers": len(customers), "books": T_books, "pages": int(T_pages),
        "symbols": int(T_sym), "sizeGB": round(T_size / 1e9, 1),
    },
    "grandSound": {k: int(v) for k, v in G_sound.items()},
    "grandPen": {k: int(v) for k, v in G_pen.items()},
    "customers": customers,
}
json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("편집 고객사", len(customers), "· 북", T_books, "· 심볼", int(T_sym), "· 리소스", round(T_size/1e9,1), "GB")
print("소리펜 모드 총합:", out["grandSound"])
print("필기펜 모드 총합:", out["grandPen"])
PY = None
